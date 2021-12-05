import json
from openpyxl import load_workbook
import re

#selenium driver
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.keys import Keys
from driver import  get_driver

#ignore warning
import sys
import warnings
if not sys.warnoptions:
    warnings.simplefilter("ignore")


def cleanStrValue(value):

    """
    arguments:   
        value : str 
    Return:             
        str : formated str
    """    
    if value:
        value = str(value).strip()
        value = value.replace(u'\u00a0', ' ')
        value = re.sub(r'[^\x00-\x7F]+', ' ', ' ' + str(value) + ' ').strip()
    else:
        value = ''
        
    return value


def load_data(inputexcel , sheetname = "dockerimages" , container_name_col = "A"):

    """
    Args:
        inputexcel:
    Returns:
        dict: dictionary of dockerimage name entity as key and url as value
    """
    
    if not inputexcel:
        return None

    sheet = load_workbook(filename = inputexcel)
    sheet_data = sheet[sheetname]
    max_row = sheet_data.max_row
    containers = {}
    for xint in range(2, max_row+1):
        row_num = str(xint)    
        container_name = cleanStrValue(value = sheet_data[container_name_col + row_num].value)
        containers[row_num] = container_name
       
    return containers


def get_containers(url , new_driver, all = False, default_container = "tomcat"):
    """
    Args:
        url: str
        new_driver: Selenium Class instance
        all: boolean
        default_container: str

    Returns:

        containers: list 
        
     """
    
    new_driver.open_driver(url)
    
    if all : 
        search_url = new_driver.all_containers_url(default_container)
        
        return search_url
    
    else:

        containers = new_driver.get_base_os()
       
        return containers


def search_base_os(entity:str) -> str:
    """
    Args:
        entity: 
    Returns:
        dict: a dictionary containing windows and linux based OS
    """
  
    base_os = {}
    url = "https://hub.docker.com/"
    new_driver = SeleniumDriver()
    search_url = get_containers(url, new_driver , all = True , default_container = entity)

    #Filter for official product and verified publisher
    filter = "&image_filter=store%2Cofficial" 
  
    windows_base_images =   search_url+filter+'&operating_system=windows'
    linux_base_images   =   search_url+filter+'&operating_system=linux'

    new_driver.driver.switch_to.new_window('tab')
    base_os["Windows base os"] =  get_containers(windows_base_images, new_driver , all = False)

    new_driver.driver.switch_to.new_window('tab')
    base_os["linux base os"] =  get_containers(linux_base_images,  new_driver,  all = False)

    new_driver.close_driver()
    return base_os


class SeleniumDriver():

    def __init__(self) -> None:

        self.driver = get_driver()

    def open_driver(self, url:str):
        """
        Args:
           url: str
        Returns:
            open the selenium driver
        """
        self.driver.get(url)

    def close_driver(self):
        """
        Args:
        
        Returns:
            close the Selenium driver
        """ 
        self.driver.quit()


    def get_all_xpath(self):
        """
        Args:
         
        Returns:
            xpath_element: 
        """
        return self.driver.find_elements(By.XPATH ,".//*" )

    def all_containers_url(self, entity:str) -> str:

        """
        Args:
            entity: name entity
        Returns:
            str: search results url
        """
      
        search_url = ""
        elem =self.get_all_xpath() 
        for el in elem:
            if el.tag_name =="input":
             #   el.clear() 
                el.send_keys(entity)
                el.send_keys(Keys.ENTER)
                search_url = self.driver.current_url
                break
        return search_url

    def find_element_by_class_name(self):
        """
        Args:
           
        Returns:
            elements containing search results
        """
        
        elements = None

        try:
            if   "No results" in self.driver.find_element(By.CLASS_NAME, "styles__searchHeader___28vtd").text:
                elements = None
            else: 
                elements = self.driver.find_element(By.CLASS_NAME , "styles__resultsWrapper___38JCx") 

        except Exception as e:
            print("There are no results for this search in Docker Hub.  ")

        return  elements

    def get_base_os(self):
        """
        Args:
        
        Returns: 
            List of base OS containers or None if no base OS if found

        """
        container_elems = self.find_element_by_class_name()
        
        if container_elems == None:
            return  "NA"
        else:

            container_elements = container_elems.find_elements_by_tag_name("a")
            containers = {}
            containers["Official image"] = []
            containers["Verified Publisher"] = []

            for cont in container_elements:

                if cont.get_attribute("data-testid") == "imageSearchResult": 
                    if len(containers) < 5:
                       
                        if "Official" in cont.text : containers["Official image"].append(cont.get_attribute("href"))
                        else: containers["Verified Publisher"].append(cont.get_attribute("href"))

                    else: break
            return  containers

def save_search_results(search_results):
    
    with open("base_os.json", "w", encoding = 'utf-8') as base_os_file:
        base_os_file.write(json.dumps(search_results, indent= 4))


def search_result(entities):

    #len entities = 305 # 1h40
    search_results = {}
    for idx ,   image in  entities.items():  
        search_results[image] = search_base_os(image)
    return search_results
    


def filter_entity(entity_names1, entity_names2):

    ent_name1=  [name1.lower() for _ , name1 in entity_names1.items()]
    ent_name2 = [name2.lower() for _, name2 in entity_names2.items()]
    ent_name = list(set(ent_name1 + ent_name2))
    
    entities = {}
    for i , name in enumerate(ent_name):
        entities[str(i)] = name
    
    return entities

if __name__ == "__main__":

    #Load data from xlxs files
    file_pth =     "kb/ACA_KG_0.0.2.0.xlsx"
    entity_names1 = load_data(file_pth , sheetname = "dockerimages" , container_name_col = "A")

    docker_images_pth = "kb/DockerHubImages.xlsx"
    entity_names2 = load_data(docker_images_pth, sheetname = "Sheet1", container_name_col  = "B")

    entity_names  = filter_entity(entity_names1, entity_names2)
    base_os_data  = search_result(entity_names)
    save_search_results(base_os_data)

    
    # #Sample single input data
    # sample_entity_name = "mysql"
    # input_data = {"1": sample_entity_name}
    # print(search_result(input_data))






    
    
        


    