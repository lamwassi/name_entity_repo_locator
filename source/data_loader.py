import json
from openpyxl import load_workbook
import re

def cleanStrValue(value):

    """
    arguments:   
        value : str 
    Return:             
        str : formated str
    """    
    if value:
        value = str(value).strip()
        value = value.replace(u'\u00a0', ' ')
        value = re.sub(r'[^\x00-\x7F]+', ' ', ' ' + str(value) + ' ').strip()
    else:
        value = ''
        
    return value

  
def load_data(inputexcel , sheetname = "dockerimages" , container_name_col = "A"):

    """
    Args:
        inputexcel:
    Returns:
        dict: dictionary of dockerimage name entity as key and url as value
    """
    
    if not inputexcel:
        return None

    sheet = load_workbook(filename = inputexcel)
    sheet_data = sheet[sheetname]
    max_row = sheet_data.max_row
    containers = {}
    for xint in range(2, max_row+1):
        row_num = str(xint)    
        container_name = cleanStrValue(value = sheet_data[container_name_col + row_num].value)
        containers[row_num] = container_name
       
    return containers


def filter_entity(entity_names1, entity_names2):
    """
    """

    ent_name1=  [name1.lower() for _ , name1 in entity_names1.items()]
    ent_name2 = [name2.lower() for _, name2 in entity_names2.items()]
    ent_name = list(set(ent_name1 + ent_name2))
    
    entities = {}
    for i , name in enumerate(ent_name):
        entities[str(i)] = name
    
    return entities